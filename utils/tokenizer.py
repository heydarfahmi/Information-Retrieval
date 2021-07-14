import csv
import openpyxl
from pathlib import Path
import joblib
from utils.normalizer import clean_text, clean_word
from utils.vectors import vectorizing_docs


def tokenize_doc(text):
    tokens = {}
    text = clean_text(text)
    words = text.split()
    for word in words:
        if 'http' in word:
            continue
        word_list = clean_word(word)
        for token in word_list:
            if token in tokens:
                tokens[token] += 1
            else:
                tokens[token] = 1

    return tokens


def find_stop_words(tokens, treshhold):
    sorted_ts = {k: v for k, v in sorted(tokens.items(), key=lambda item: item[1])}
    for t in sorted_ts:
        if treshhold == 0:
            break
        treshhold -= 1


def make_champion(tokens, champion_list_size):
    champion_list = {}
    for token in tokens:
        sorted_docs = list(dict(sorted(tokens[token].items(), key=lambda item: item[1], reverse=True)).keys())
        treshold = min(champion_list_size, len(sorted_docs))
        champion_list[token] = sorted_docs[0:treshold]
    return champion_list


def tokenize_documents(documents, vectorizing=False, champion_list_size=0):
    tokens = {}
    docs_norms = {}
    for doc_id, text in enumerate(documents, 1):
        docs_norms[doc_id] = 0
        doc_tokens = tokenize_doc(text)
        for token in doc_tokens:
            if token in tokens:

                tokens[token].update({doc_id: doc_tokens[token]})
            else:
                tokens[token] = {doc_id: doc_tokens[token]}
    # # if we need counts
    print("done")
    ls = {token: sum(tokens[token].values()) for token in tokens.keys()}
    find_stop_words(ls, 20)  # print this then save
    if vectorizing:
        champion_list = {}
        if champion_list_size:
            champion_list = make_champion(tokens, champion_list_size)
        return vectorizing_docs(tokens, len(documents), docs_norms), champion_list

    return {token: list(tokens[token].keys()) for token in tokens.keys()}


class Tokenizer:
    def __init__(self, data_set, post_index_path=None, champion_list_size=10, update_post=False, vectorizing=False):
        self.data_set = data_set
        self.post_index = post_index_path
        self.urls = []
        self.N = 0
        self.tokens = {}
        self.champion_list = {}
        post_index_file = Path('./', post_index_path)
        if post_index_file.is_file() and not update_post:
            self.load(post_index_path)
        else:
            update_post = True
        if update_post:

            documents = self.get_documents()
            self.tokenize(documents, vectorizing, champion_list_size=champion_list_size)
            self.save(post_index_path)

    def tokenize(self, docs, make_vec, champion_list_size):
        if make_vec:
            self.tokens, self.champion_list = tokenize_documents(docs, make_vec, champion_list_size)
        else:
            self.tokens = tokenize_documents(docs, make_vec, champion_list_size)

    def load(self, post_index_path):
        self.tokens = joblib.load(post_index_path)
        self.champion_list = joblib.load("champion_list.pkl")
        self.urls = joblib.load("url.pkl")
        self.N = len(self.urls)

    def save(self, post_index_path):
        joblib.dump(self.tokens, post_index_path)
        joblib.dump(self.urls, "url.pkl")
        self.champion_list = joblib.dump(self.champion_list, "champion_list.pkl")

    def get_documents(self):
        filename = self.data_set
        xlsx_file = Path('utils/dataset', filename)
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj.active

        documents = []
        for row in sheet.iter_rows(2, sheet.max_row):
            documents.append(row[1].value)
            self.urls.append(row[2].value)
        self.N = len(self.urls)
        return documents
